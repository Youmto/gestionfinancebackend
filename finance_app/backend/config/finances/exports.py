"""
Services d'export pour les finances (PDF, Excel)
"""

import io
from decimal import Decimal
from datetime import datetime
from typing import List, Dict

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm, mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont


class ExcelExporter:
    """Exportateur de données en format Excel"""
    
    # Styles
    HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    
    INCOME_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    EXPENSE_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    @classmethod
    def export_transactions(cls, transactions, user, title="Transactions") -> HttpResponse:
        """
        Exporte les transactions en Excel.
        
        Args:
            transactions: QuerySet de transactions
            user: Utilisateur
            title: Titre du rapport
            
        Returns:
            HttpResponse avec le fichier Excel
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"
        
        # En-tête du rapport
        ws.merge_cells('A1:G1')
        ws['A1'] = f"Rapport: {title}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:G2')
        ws['A2'] = f"Généré le {timezone.now().strftime('%d/%m/%Y à %H:%M')} pour {user.email}"
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # En-têtes des colonnes
        headers = ['Date', 'Type', 'Catégorie', 'Description', 'Montant', 'Devise', 'Groupe']
        row_num = 4
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.fill = cls.HEADER_FILL
            cell.font = cls.HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border = cls.BORDER
        
        # Données
        total_income = Decimal('0')
        total_expense = Decimal('0')
        
        for transaction in transactions:
            row_num += 1
            
            row_data = [
                transaction.date.strftime('%d/%m/%Y'),
                'Revenu' if transaction.type == 'income' else 'Dépense',
                transaction.category.name if transaction.category else '-',
                transaction.description or '-',
                float(transaction.amount),
                'XAF',
                transaction.group.name if transaction.group else '-'
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = cls.BORDER
                
                if col_num == 5:  # Colonne montant
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
            
            # Colorer selon le type
            fill = cls.INCOME_FILL if transaction.type == 'income' else cls.EXPENSE_FILL
            for col_num in range(1, 8):
                ws.cell(row=row_num, column=col_num).fill = fill
            
            # Totaux
            if transaction.type == 'income':
                total_income += transaction.amount
            else:
                total_expense += transaction.amount
        
        # Résumé
        row_num += 2
        ws.cell(row=row_num, column=4, value="Total Revenus:").font = Font(bold=True)
        ws.cell(row=row_num, column=5, value=float(total_income)).number_format = '#,##0.00'
        ws.cell(row=row_num, column=5).fill = cls.INCOME_FILL
        
        row_num += 1
        ws.cell(row=row_num, column=4, value="Total Dépenses:").font = Font(bold=True)
        ws.cell(row=row_num, column=5, value=float(total_expense)).number_format = '#,##0.00'
        ws.cell(row=row_num, column=5).fill = cls.EXPENSE_FILL
        
        row_num += 1
        ws.cell(row=row_num, column=4, value="Solde:").font = Font(bold=True)
        balance = total_income - total_expense
        ws.cell(row=row_num, column=5, value=float(balance)).number_format = '#,##0.00'
        ws.cell(row=row_num, column=5).fill = cls.INCOME_FILL if balance >= 0 else cls.EXPENSE_FILL
        
        # Ajuster la largeur des colonnes
        column_widths = [12, 12, 20, 35, 15, 8, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Créer la réponse
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"transactions_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    @classmethod
    def export_budget_report(cls, categories, user, year, month) -> HttpResponse:
        """Exporte le rapport de budget en Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Budget"
        
        month_names = ['Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
                       'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']
        
        # En-tête
        ws.merge_cells('A1:F1')
        ws['A1'] = f"Rapport Budget - {month_names[month-1]} {year}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:F2')
        ws['A2'] = f"Généré le {timezone.now().strftime('%d/%m/%Y')} pour {user.email}"
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # En-têtes
        headers = ['Catégorie', 'Budget', 'Dépensé', 'Restant', '%', 'Statut']
        row_num = 4
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.fill = cls.HEADER_FILL
            cell.font = cls.HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border = cls.BORDER
        
        # Données
        for category in categories:
            status = category.get_budget_status(year, month)
            if not status:
                continue
            
            row_num += 1
            
            # Déterminer le statut texte
            if status['is_over_budget']:
                status_text = '🚨 Dépassé'
            elif status['is_alert']:
                status_text = '⚠️ Attention'
            else:
                status_text = '✅ OK'
            
            row_data = [
                f"{category.icon} {category.name}",
                float(status['budget']),
                float(status['spent']),
                float(status['remaining']),
                f"{status['percentage']:.1f}%",
                status_text
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = cls.BORDER
                
                if col_num in [2, 3, 4]:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
            
            # Colorer selon le statut
            if status['is_over_budget']:
                fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            elif status['is_alert']:
                fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            else:
                fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
            
            for col_num in range(1, 7):
                ws.cell(row=row_num, column=col_num).fill = fill
        
        # Largeurs
        column_widths = [25, 15, 15, 15, 10, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Réponse
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"budget_{year}_{month:02d}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response


class PDFExporter:
    """Exportateur de données en format PDF"""
    
    @classmethod
    def export_transactions(cls, transactions, user, title="Transactions") -> HttpResponse:
        """Exporte les transactions en PDF."""
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=1.5*cm,
            leftMargin=1.5*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Titre
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=12,
            alignment=1  # Centre
        )
        elements.append(Paragraph(title, title_style))
        
        # Sous-titre
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            alignment=1
        )
        elements.append(Paragraph(
            f"Généré le {timezone.now().strftime('%d/%m/%Y à %H:%M')} pour {user.email}",
            subtitle_style
        ))
        elements.append(Spacer(1, 20))
        
        # Tableau des transactions
        data = [['Date', 'Type', 'Catégorie', 'Description', 'Montant']]
        
        total_income = Decimal('0')
        total_expense = Decimal('0')
        
        for t in transactions:
            amount_str = f"{t.amount:,.0f} XAF"
            if t.type == 'expense':
                amount_str = f"-{amount_str}"
            
            data.append([
                t.date.strftime('%d/%m/%Y'),
                'Revenu' if t.type == 'income' else 'Dépense',
                t.category.name if t.category else '-',
                (t.description[:30] + '...') if t.description and len(t.description) > 30 else (t.description or '-'),
                amount_str
            ])
            
            if t.type == 'income':
                total_income += t.amount
            else:
                total_expense += t.amount
        
        # Ajouter les totaux
        data.append(['', '', '', 'Total Revenus:', f"{total_income:,.0f} XAF"])
        data.append(['', '', '', 'Total Dépenses:', f"-{total_expense:,.0f} XAF"])
        data.append(['', '', '', 'Solde:', f"{total_income - total_expense:,.0f} XAF"])
        
        # Style du tableau
        table = Table(data, colWidths=[2.5*cm, 2.5*cm, 3.5*cm, 5*cm, 3*cm])
        
        table_style = TableStyle([
            # En-tête
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            
            # Corps
            ('FONTNAME', (0, 1), (-1, -4), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (4, 1), (4, -1), 'RIGHT'),
            
            # Totaux
            ('FONTNAME', (3, -3), (-1, -1), 'Helvetica-Bold'),
            ('LINEABOVE', (3, -3), (-1, -3), 1, colors.black),
            
            # Bordures
            ('GRID', (0, 0), (-1, -4), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ])
        
        # Colorer les lignes selon le type
        for i, t in enumerate(transactions, 1):
            if t.type == 'income':
                table_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#D1FAE5'))
            else:
                table_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#FEE2E2'))
        
        table.setStyle(table_style)
        elements.append(table)
        
        # Générer le PDF
        doc.build(elements)
        
        buffer.seek(0)
        filename = f"transactions_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    @classmethod
    def export_monthly_report(cls, user, year, month, transactions, categories) -> HttpResponse:
        """Génère un rapport mensuel complet en PDF."""
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=1.5*cm,
            leftMargin=1.5*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        month_names = ['Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
                       'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']
        
        # Titre
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            spaceAfter=6,
            alignment=1
        )
        elements.append(Paragraph(f"Rapport Financier", title_style))
        elements.append(Paragraph(
            f"{month_names[month-1]} {year}",
            ParagraphStyle('MonthTitle', parent=styles['Heading2'], alignment=1, textColor=colors.HexColor('#4F46E5'))
        ))
        elements.append(Spacer(1, 10))
        elements.append(Paragraph(
            f"Pour: {user.first_name} {user.last_name} ({user.email})",
            ParagraphStyle('UserInfo', parent=styles['Normal'], alignment=1, textColor=colors.grey)
        ))
        elements.append(Spacer(1, 30))
        
        # Résumé
        total_income = sum(t.amount for t in transactions if t.type == 'income')
        total_expense = sum(t.amount for t in transactions if t.type == 'expense')
        balance = total_income - total_expense
        
        summary_data = [
            ['Résumé Financier', ''],
            ['Total Revenus', f"{total_income:,.0f} XAF"],
            ['Total Dépenses', f"{total_expense:,.0f} XAF"],
            ['Solde', f"{balance:,.0f} XAF"],
            ['Nombre de transactions', str(len(transactions))]
        ]
        
        summary_table = Table(summary_data, colWidths=[8*cm, 5*cm])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('SPAN', (0, 0), (-1, 0)),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 30))
        
        # Budget
        elements.append(Paragraph("Suivi des Budgets", styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        budget_data = [['Catégorie', 'Budget', 'Dépensé', 'Restant', '%']]
        
        for cat in categories:
            status = cat.get_budget_status(year, month)
            if status:
                budget_data.append([
                    f"{cat.icon} {cat.name}",
                    f"{status['budget']:,.0f}",
                    f"{status['spent']:,.0f}",
                    f"{status['remaining']:,.0f}",
                    f"{status['percentage']:.0f}%"
                ])
        
        if len(budget_data) > 1:
            budget_table = Table(budget_data, colWidths=[5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2*cm])
            budget_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(budget_table)
        else:
            elements.append(Paragraph("Aucun budget défini.", styles['Normal']))
        
        # Générer
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"rapport_{year}_{month:02d}.pdf"
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response